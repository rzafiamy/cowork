import asyncio
import os
import sys
import shutil
import click
from pathlib import Path
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich import box
import pandas as pd

# Add CLI dir to sys path for imports
project_root = Path(__file__).parent.parent
cli_dir = project_root / "cli"
sys.path.insert(0, str(cli_dir))

from cowork.agent import GeneralPurposeAgent
from cowork.config import ConfigManager, Session, Scratchpad, JobManager, AgentJob
from cowork.memoria import Memoria
from cowork.api_client import APIClient

console = Console()

class EvalCore:
    def __init__(self):
        self.config = ConfigManager()

    async def run_test_suite(self, test_cases):
        console.print(Panel("[bold primary]🚀 Starting Cowork Agentic Evaluation Suite[/bold primary]", border_style="cyan"))
        
        results = []
        
        table = Table(box=box.MINIMAL_DOUBLE_HEAD, show_lines=True)
        table.add_column("Category", style="cyan")
        table.add_column("Test Case", style="white")
        table.add_column("Status", justify="center")
        
        for tc in test_cases:
            with console.status(f"[yellow]Evaluating:[/yellow] {tc['name']}..."):
                session = Session(title=f"Test Session - {tc['name']}")
                
                api_client = APIClient(
                    endpoint=self.config.api_endpoint,
                    api_key=self.config.api_key,
                    token_callback=lambda m, u: None
                )
                job_mgr = JobManager(max_jobs=10)
                scratchpad = Scratchpad(session.session_id)
                memoria = Memoria("eval_test_user", session.session_id, api_client, self.config)
                
                agent = GeneralPurposeAgent(
                    api_client=api_client,
                    config=self.config,
                    scratchpad=scratchpad,
                    memoria=memoria,
                    job_manager=job_mgr,
                    status_callback=lambda x: None,
                    stream_callback=lambda x: None,
                )
                
                async def auto_confirm(name, reason, args):
                    return True
                agent.confirm_cb = auto_confirm

                final_response = ""
                context = {}
                for prompt in tc["prompts"]:
                    job = AgentJob(session_id=session.session_id, prompt=prompt)
                    job_mgr.register(job)
                    job_mgr.start(job.job_id)
                    final_response = await agent.run(prompt, session, job)
                    session.add_message("user", prompt)
                    session.add_message("assistant", final_response)
                
                passed = False
                try:
                    passed = tc["verify"](final_response, context)
                except Exception as e:
                    passed = False
                    
                status_text = "[bold green]PASS ✅[/bold green]" if passed else "[bold red]FAIL ❌[/bold red]"
                table.add_row(tc['category'], tc['name'], status_text)
                
                results.append({
                    "Category": tc['category'],
                    "Test Name": tc['name'],
                    "Prompts": " -> ".join(tc['prompts']),
                    "Last Response Snippet": str(final_response)[:200] + ("..." if len(str(final_response)) > 200 else ""),
                    "Status": "PASSED" if passed else "FAILED"
                })
                
                # Cleanup file test artifacts
                if tc["name"] == "Create Directory & File":
                    if Path("test_agent_dir").exists():
                        shutil.rmtree("test_agent_dir")
                        
        console.print(table)
        return results

    def generate_excel_report(self, results, filename="evaluation_report.xlsx"):
        df = pd.DataFrame(results)
        
        writer = pd.ExcelWriter(filename, engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Eval Results")
        
        workbook = writer.book
        worksheet = writer.sheets["Eval Results"]
        from openpyxl.styles import PatternFill, Font, Alignment
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        pass_font = Font(color="006100")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(color="9C0006")
        
        status_col_idx = 5
        for row in worksheet.iter_rows(min_row=2):
            cell = row[status_col_idx-1]
            if cell.value == "PASSED":
                cell.fill = pass_fill
                cell.font = pass_font
            else:
                cell.fill = fail_fill
                cell.font = fail_font
                
            row[2].alignment = Alignment(wrap_text=True)
            row[3].alignment = Alignment(wrap_text=True)
                
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 50
        worksheet.column_dimensions['D'].width = 60
        worksheet.column_dimensions['E'].width = 15
        
        writer.close()
        console.print(f"[bold green]📊 Excel report saved to tests/{filename}[/bold green]")
